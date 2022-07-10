import openpyxl
from openpyxl.styles import NamedStyle, Font, Side, Border, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from forecasting import forecasting


def create_report():
    df, analysis_dict, forecasting_period, orders = forecasting()

    # Create new excel book
    report_book = openpyxl.Workbook()
    report_book.remove(report_book.active)
    report_book.create_sheet(title='Report', index=0)
    report_sheet = report_book['Report']

    # Create styles
    ns_header = NamedStyle(name='header')
    ns_header.font = Font(name='Arial', bold=True, size=8)
    ns_header.fill = PatternFill('solid', fgColor='DDDDDD')
    thin = Side(border_style='thin', color='000000')
    ns_header.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ns_header.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    report_book.add_named_style(ns_header)

    ns_info = NamedStyle(name='info')
    thin = Side(border_style='dotted', color='000000')
    ns_info.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ns_info.font = Font(name='Arial', bold=False, size=8)
    report_book.add_named_style(ns_info)

    # Create sheet header
    column_names = df.columns
    start_position = 3
    for index, name in enumerate(column_names):
        cell = report_sheet.cell(row=start_position, column=index+1)
        info = name
        cell.value = info
        cell.style = 'header'
        if 'cost' in info or 'Расх.' in info:
            cell.fill = PatternFill('solid', fgColor='CCECFF')
        elif 'sale' in info or 'Прод.' in info:
            cell.fill = PatternFill('solid', fgColor='CCFFCC')
        elif 'stock' in info:
            cell.fill = PatternFill('solid', fgColor='CCC0DA')
    for col in range(1, df.shape[1]+1):
        cell = report_sheet.cell(row=3, column=col)
        info = cell.value
        if 'cost ' in info or 'sale ' in info:
            cell.value = info[5:]
        elif 'stock ' in info:
            cell.value = info[6:]

    # Create a dictionary to adjust column widths
    column_width_dict = {}
    max_col = df.shape[1]
    for col in range(1, max_col+1):
        if col == 1:
            column_width_dict[col] = 50
        elif 2 <= col <= 4:
            column_width_dict[col] = 8
        elif orders > 0:
            if 5 <= col <= 4 + orders:
                column_width_dict[col] = 12
            else:
                column_width_dict[col] = 6
        else:
            column_width_dict[col] = 10

    # Setting column widths
    for k, v in column_width_dict.items():
        report_sheet.column_dimensions[get_column_letter(k)].width = v

    # Create a tuple that will be needed to group columns when creating a report
    column_dimensions_group = (
        (2, 3),
        [(), (5, 4 + orders)][orders > 0],
        (4 + orders + 4 + 3 + 1, 4 + orders + 4 + 3 + forecasting_period),
        (4 + orders + 4 + 3 + forecasting_period + 3 + 1, 4 + orders + 4 + 3 + forecasting_period*2 + 3)
    )

    # Column grouping and header freezing
    report_sheet.freeze_panes = report_sheet.cell(row=4, column=2)
    for item in column_dimensions_group:
        if item:
            report_sheet.column_dimensions.group(get_column_letter(item[0]),
                                                 get_column_letter(item[-1]),
                                                 hidden=True)

    # Transfer of information from df
    start_row = 4
    for row in range(start_row, df.shape[0]+start_row):
        for col in range(1, df.shape[1]+1):
            info = df.iloc[row-start_row][col-1]
            cell = report_sheet.cell(row=row, column=col)
            cell.value = info
            cell.style = 'info'

    # Painting cells depending on the stock
    stock_start_col = 4 + orders + 4 + forecasting_period*2 + 6 + 1
    for row in range(start_row, df.shape[0]+start_row):
        for col in range(stock_start_col, df.shape[1]):
            cell_1 = report_sheet.cell(row=row, column=1)
            cell_2 = report_sheet.cell(row=row, column=col)

            if 'r' in analysis_dict[str(cell_1.value)][4][col-stock_start_col]:
                cell_2.fill = PatternFill('solid', fgColor='FFC7CE')
            elif 'y' in analysis_dict[str(cell_1.value)][4][col-stock_start_col]:
                cell_2.fill = PatternFill('solid', fgColor='FFFFCC')
            elif 'g' in analysis_dict[str(cell_1.value)][4][col-stock_start_col]:
                cell_2.fill = PatternFill('solid', fgColor='C6EFCE')

    report_book.save(f'data/!Report.xlsx')
    print('Отчет создан!')


def main():
    create_report()


if __name__ == '__main__':
    main()
